from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

NAVY='1A1A2E'; AMBER='E8C547'; WHITE='FFFFFF'; LIGHT='F2F2F2'
MGREY='DDDDDD'; DGREY='BBBBBB'; AMBER_BG='FFF9E0'; RED_F='C00000'
SECTION_BG='1A1A2E'; SECTION_F='FFFFFF'; SUBTOT_BG='FFF2CC'

def F(bold=False,color='000000',size=9,italic=False):
    return Font(name='Arial',bold=bold,color=color,size=size,italic=italic)
def fill(c): return PatternFill('solid',fgColor=c)
def al(h='left',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def bdr(col=MGREY):
    s=Side(style='thin',color=col)
    return Border(left=s,right=s,top=s,bottom=s)
def thick_bdr():
    s=Side(style='medium',color=NAVY)
    return Border(left=s,right=s,top=s,bottom=s)

GBP='£#,##0.00'; INT='#,##0'

def cell(ws,r,c,v=None,bg=WHITE,fc='000000',bold=False,h='left',fmt=None,sz=9,italic=False,wrap=False):
    cl=ws.cell(r,c)
    if v is not None: cl.value=v
    cl.font=Font(name='Arial',bold=bold,color=fc,size=sz,italic=italic)
    cl.fill=fill(bg)
    cl.alignment=al(h,'center',wrap)
    cl.border=bdr()
    if fmt: cl.number_format=fmt
    return cl

def sect(ws,r,c1,c2,txt):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cl=ws.cell(r,c1)
    cl.value=txt; cl.font=F(True,SECTION_F,9); cl.fill=fill(SECTION_BG)
    cl.alignment=al('left','center'); ws.row_dimensions[r].height=14

def tot(ws,r,c,formula,fmt=GBP):
    cl=ws.cell(r,c)
    cl.value=formula; cl.font=F(True,NAVY,9); cl.fill=fill(AMBER)
    cl.alignment=al('right','center'); cl.border=thick_bdr(); cl.number_format=fmt

# ── Scheme data ───────────────────────────────────────────────────────────────
STUDIO=770; ACC_N=43; TOTAL=813
S_TOTAL=4290.50; A_TOTAL=5248.89
BED_T=STUDIO*1511.07+ACC_N*1734.34; KIT_T=STUDIO*1331.80+ACC_N*2182.68
LOOSE_T=STUDIO*574.52+ACC_N*574.52; APP_T=STUDIO*873.11+ACC_N*757.34
PM_T=TOTAL*68.43; SUB_T=BED_T+KIT_T+LOOSE_T+APP_T+PM_T
MCD_T=SUB_T*0.025; GRAND=SUB_T-MCD_T

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1: SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
ws=wb.active; ws.title='SUMMARY'
ws.column_dimensions['A'].width=3
ws.column_dimensions['B'].width=42
ws.column_dimensions['C'].width=18
ws.column_dimensions['D'].width=4

# Header
ws.merge_cells('B1:C1')
cl=ws['B1']; cl.value='FLAGSTAFFE  —  FF&E Quotation'
cl.font=Font(name='Arial',bold=True,size=14,color=WHITE)
cl.fill=fill(NAVY); cl.alignment=al('center','center')
ws.row_dimensions[1].height=28

# Project info block
r=3
info=[
    ('Project:','Heavitree Road, Exeter — NCO (Seven) Ltd / Nixon Homes Ltd'),
    ('Planning reference:','25/0676/FUL — Exeter City Council'),
    ('Issued to:','[Contact TBC]'),
    ('Date:','2 June 2026'),
    ('Revision:','0'),
    ('Sales:','Dan Brownsword  07442399616'),
    ('Estimator:','Jimmy Hogg  07392765933'),
    ('Flag reference:','Flag-HVR-001'),
    ('Specification:','Kronospan Band A  |  18mm carcass  |  Metabox soft close'),
    ('Inflation basis:','January 2027 start (factor 1.0527)'),
    ('Programme:','8 months, even and continuous delivery'),
]
for label,val in info:
    cell(ws,r,2,label,LIGHT,NAVY,True,'left',sz=9)
    cell(ws,r,3,val,WHITE,'000000',False,'left',sz=9)
    ws.row_dimensions[r].height=14; r+=1

r+=1

# Room counts
sect(ws,r,2,3,'ROOM COUNTS'); r+=1
cell(ws,r,2,'Standard Studios',LIGHT,NAVY,True,'left',sz=9)
cell(ws,r,3,STUDIO,WHITE,'000000',False,'right',INT,9)
ws.row_dimensions[r].height=14; r+=1
cell(ws,r,2,'DDA/Accessible Studios',LIGHT,NAVY,True,'left',sz=9)
cell(ws,r,3,ACC_N,WHITE,'000000',False,'right',INT,9)
ws.row_dimensions[r].height=14; r+=1
cell(ws,r,2,'Total units',LIGHT,NAVY,True,'left',sz=9)
cell(ws,r,3,TOTAL,SUBTOT_BG,NAVY,True,'right',INT,9)
ws.row_dimensions[r].height=14; r+=2

# Cost breakdown
sect(ws,r,2,3,'COST BREAKDOWN'); r+=1
costs=[('Bedrooms',BED_T),('Kitchens',KIT_T),('Loose Goods',LOOSE_T),('Appliances',APP_T)]
cost_rows=[]
for label,val in costs:
    cell(ws,r,2,label,WHITE,'000000',False,'left',sz=9)
    cell(ws,r,3,val,WHITE,'000000',False,'right',GBP,9)
    cost_rows.append(r); ws.row_dimensions[r].height=15; r+=1

# PM fee
cell(ws,r,2,'Project Management Fee',WHITE,'000000',False,'left',sz=9)
cell(ws,r,3,PM_T,WHITE,'000000',False,'right',GBP,9)
pm_row=r; ws.row_dimensions[r].height=15; r+=1

# Sub-total
cell(ws,r,2,'Sub-total',SUBTOT_BG,NAVY,True,'left',sz=10)
cell(ws,r,3,SUB_T,SUBTOT_BG,NAVY,True,'right',GBP,10)
sub_row=r; ws.row_dimensions[r].height=16; r+=1

# MCD
cell(ws,r,2,'2.5% MCD Discount',LIGHT,RED_F,True,'left',sz=9)
mcd_c=ws.cell(r,3); mcd_c.value=-MCD_T
mcd_c.font=Font(name='Arial',bold=True,color=RED_F,size=9)
mcd_c.fill=fill(LIGHT); mcd_c.alignment=al('right','center')
mcd_c.border=bdr(); mcd_c.number_format=GBP
ws.row_dimensions[r].height=15; r+=1

# TOTAL
cell(ws,r,2,'TOTAL (excl. VAT)',NAVY,WHITE,True,'left',sz=12)
tc=ws.cell(r,3); tc.value=GRAND
tc.font=Font(name='Arial',bold=True,color=NAVY,size=12)
tc.fill=fill(AMBER); tc.alignment=al('right','center')
tc.border=thick_bdr(); tc.number_format=GBP
ws.row_dimensions[r].height=22; r+=2

# Notes
notes=[
    'excl. VAT — Reverse Charge is expected to apply',
    'Inflation built in for commencement on site: January 2027',
    f'Priced for even and continuous delivery over: 8 months',
    'Pricing based on standard terms — see T&Cs sheet',
    'Items noted as RATE ONLY do not form part of this quotation',
    'Finishes priced: Kronospan Band A',
]
for i,n in enumerate(notes,1):
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
    cell(ws,r,2,f'{i}  {n}',WHITE,'555555',False,'left',sz=8,italic=True)
    ws.row_dimensions[r].height=13; r+=1

r+=1
ws.merge_cells(f'B{r}:C{r}')
cl=ws[f'B{r}']
cl.value='enquiries@flagstaffe.com  |  01260 460 003  |  www.flagstaffe.com  |  CRN: 14154708  |  VAT: 421442927'
cl.font=Font(name='Arial',size=8,color=WHITE)
cl.fill=fill(NAVY); cl.alignment=al('center','center')
ws.row_dimensions[r].height=14

# ─────────────────────────────────────────────────────────────────────────────
# Helper: build a detail sheet
# ─────────────────────────────────────────────────────────────────────────────
def build_detail(wb, title, items, s_count, a_count):
    """
    items: list of (code, description, spec, supplier, model, rate, s_qty, a_qty)
    """
    ws=wb.create_sheet(title)
    # Column widths
    widths={'A':3,'B':6,'C':34,'D':18,'E':13,'F':12,'G':8,'H':8,'I':8,'J':13,'K':13,'L':13,'M':4}
    for col,w in widths.items(): ws.column_dimensions[col].width=w

    # Header
    ws.merge_cells('B1:L1')
    cl=ws['B1']
    cl.value=f'FLAGSTAFFE  —  {title}  |  Heavitree Road, Exeter  |  Flag-HVR-001 Rev 0'
    cl.font=Font(name='Arial',bold=True,size=11,color=WHITE)
    cl.fill=fill(NAVY); cl.alignment=al('center','center')
    ws.row_dimensions[1].height=20

    # Spec row
    ws.merge_cells('B2:L2')
    cl=ws['B2']
    cl.value='Décor: Kronospan Band A  |  18mm carcass  |  Metabox soft close drawers  |  155° soft-close hinges  |  Discreet handle'
    cl.font=Font(name='Arial',size=8,italic=True,color='555555')
    cl.fill=fill(LIGHT); cl.alignment=al('center','center')
    ws.row_dimensions[2].height=12

    # Column headers row 3
    headers=[
        (2,'Code','left'),(3,'Description','left'),(4,'Spec','left'),
        (5,'Supplier','left'),(6,'Model','left'),(7,'Rate (£)','right'),
        (8,f'Studio\nqty','center'),(9,f'Studio\nsub-total','right'),
        (10,f'ACC\nqty','center'),(11,f'ACC\nsub-total','right'),
        (12,'TOTAL (£)','right'),
    ]
    ws.row_dimensions[3].height=28
    for col,label,h in headers:
        cl=ws.cell(3,col)
        cl.value=label
        cl.font=Font(name='Arial',bold=True,color=WHITE,size=9)
        cl.fill=fill(NAVY); cl.alignment=al(h,'center',True)
        cl.border=bdr(NAVY)

    # Room count reference row 4
    ws.row_dimensions[4].height=13
    ws.merge_cells('B4:G4')
    cell(ws,4,2,'Room counts (from SUMMARY sheet):',LIGHT,'555555',False,'left',sz=8,italic=True)
    cell(ws,4,8,s_count,LIGHT,'0070C0',True,'center',INT,8)
    cell(ws,4,9,'',LIGHT,'555555',False,'center',sz=8)
    cell(ws,4,10,a_count,LIGHT,'0070C0',True,'center',INT,8)
    cell(ws,4,11,'',LIGHT,'555555',False,'center',sz=8)
    cell(ws,4,12,'',LIGHT,'555555',False,'center',sz=8)

    r=5; running=0; current_section=''
    data_rows=[]

    for item in items:
        code,desc,spec,sup,model,rate,s_qty,a_qty=item

        # Section header
        section=code.split('_')[0] if '_' in code else ''
        # Use desc to detect section from a special marker
        if desc.startswith('__SECTION__'):
            s_name=desc.replace('__SECTION__','')
            ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=12)
            cl=ws.cell(r,2); cl.value=s_name
            cl.font=Font(name='Arial',bold=True,size=9,color='1A1A2E')
            cl.fill=fill(AMBER_BG); cl.alignment=al('left','center')
            ws.row_dimensions[r].height=14; r+=1; continue

        s_sub=rate*s_qty*s_count
        a_sub=rate*a_qty*a_count
        total=s_sub+a_sub
        running+=total

        ws.row_dimensions[r].height=14
        cell(ws,r,2,code,WHITE,'888888',False,'left',sz=8)
        cell(ws,r,3,desc,WHITE,'000000',False,'left',sz=9)
        cell(ws,r,4,spec,WHITE,'555555',False,'left',sz=8,italic=True,wrap=True)
        cell(ws,r,5,sup,WHITE,'555555',False,'left',sz=8)
        cell(ws,r,6,model,WHITE,'555555',False,'left',sz=8)
        cell(ws,r,7,rate,WHITE,'000000',False,'right',GBP,9)

        if s_qty:
            cell(ws,r,8,s_qty,WHITE,'000000',False,'center',INT,9)
            cell(ws,r,9,s_sub,WHITE,'000000',False,'right',GBP,9)
        else:
            cell(ws,r,8,'',LIGHT,'BBBBBB',False,'center',sz=9)
            cell(ws,r,9,'',LIGHT,'BBBBBB',False,'right',sz=9)
        if a_qty:
            cell(ws,r,10,a_qty,WHITE,'000000',False,'center',INT,9)
            cell(ws,r,11,a_sub,WHITE,'000000',False,'right',GBP,9)
        else:
            cell(ws,r,10,'',LIGHT,'BBBBBB',False,'center',sz=9)
            cell(ws,r,11,'',LIGHT,'BBBBBB',False,'right',sz=9)

        cell(ws,r,12,total,WHITE,'000000',False,'right',GBP,9)
        data_rows.append(r); r+=1

    # Total row
    ws.row_dimensions[r].height=18
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=11)
    cl=ws.cell(r,2); cl.value=f'TOTAL  —  {title}'
    cl.font=Font(name='Arial',bold=True,size=10,color=WHITE)
    cl.fill=fill(NAVY); cl.alignment=al('left','center')
    tot(ws,r,12,running)

    return ws, running

# ─────────────────────────────────────────────────────────────────────────────
# Line item data (code, desc, spec, sup, model, rate, s_qty, a_qty)
# ─────────────────────────────────────────────────────────────────────────────
SEC=lambda name: ('SEC',f'__SECTION__{name}','','','',0,0,0)

BEDROOM_ITEMS=[
  SEC('DESKS'),
  ('B1','Desktop straight 1800 x 600 x 25mm','1x cable tidy 80mm, supporting batten','Factory','25 pricelist',56.67,0,1),
  ('B4','Desktop straight 2100 x 600 x 25mm','1x cable tidy 80mm, supporting batten','Factory','25 pricelist',73.90,1,0),
  ('B6','End Panel 25mm MFC end of desk','','Factory','25 pricelist',21.77,1,0),
  ('B10','Desk Rise and Fall system — Hand Crank','DDA rooms only','RACK','',368.33,0,1),
  ('B13','Drawers 400 x 725 x 570mm — 3 drawers','','Factory','25 pricelist',108.53,1,0),
  ('B13','ACC Pedestal — 3 drawers on castors','DDA rooms only','Factory','25 pricelist',120.95,0,1),
  SEC('BEDS'),
  ('B28','Bed Double — 1420 x 1950 x 450mm','MFC bed with storage, open side access','Factory','25 pricelist',207.11,1,1),
  ('B46','Headboard Flat Double — 1400 x 550 x 18mm','','Factory','25 pricelist',28.17,1,1),
  ('B43','Extended Headboard Boxed 1650 x 1000 x 150mm','Service zone from kitchen — Standard Studio','Factory','25 pricelist',86.16,1,0),
  SEC('WARDROBES'),
  ('B61','Wardrobe 700 x 2100 x 600mm','2 doors, full shelf, full hanging rail','Station Rd','25 pricelist',244.48,1,0),
  ('B64','Wardrobe 400 x 2100 x 600mm','1 door, full shelf, full hanging rail','Station Rd','25 pricelist',208.47,1,0),
  ('B55','Wardrobe 1000 x 2100 x 600mm','2 doors, full shelf, full hanging rail','Factory','25 pricelist',284.99,0,1),
  ('B73','Tall Open Shelving 400 x 2100 x 600mm','5 fixed shelves','Factory','25 pricelist',124.27,1,1),
  ('B76','Pull-down DDA rail within wardrobe','Grey — DDA rooms','Hafele','805.20.056',95.38,0,1),
  SEC('OTHER'),
  ('B82','Coathooks on pattress','','Umbra','',14.36,1,1),
  ('B64','Tall infill 2100 x 70 x 18mm','','','',13.26,1,1),
  ('B84','Mirror 500 x 1600 x 4mm','Chrome fixings, safety backed, polished edges','Tradesliders','',46.31,1,1),
  ('B84','Mirror back panel','','','',26.92,1,1),
  ('B93','Pinboard 1800 x 600 x 9mm','Polycolour, mechanically fixed, unframed','Pitts','Polycolour',48.46,0,1),
  ('B96','Pinboard 2100 x 600 x 9mm','Polycolour, mechanically fixed, unframed','Pitts','Polycolour',64.62,1,0),
  SEC('INSTALLATION (inc. offload, distribution, unpacking, assembly, fixing, waste, snagging)'),
  ('INS','Installation — Standard Studio','Per studio bedroom','','',242.75,1,0),
  ('INS','Installation — Studio ACC/DDA','Per accessible studio','','',299.16,0,1),
]

KITCHEN_ITEMS=[
  SEC('WALL UNITS'),
  ('K4','Wall Unit 600 x 720 x 340mm — 1 door and shelf','Studios and ACC — 2 per kitchen','Factory','25 pricelist',67.85,2,2),
  ('K9','Extractor Unit 600 x 720 x 340mm — top opening with shelf','All kitchens','Factory','25 pricelist',74.48,1,1),
  SEC('BASE UNITS'),
  ('K13','Base Unit 600 x 720 x 600mm highline','ACC only','Factory','25 pricelist',85.25,0,1),
  ('K25','Sink Unit 600 x 720 x 600mm inc removable back','Standard Studio only','Factory','25 pricelist',79.03,1,0),
  ('K25','Undercounter Fridge Door 600 x 720 x 18mm on plinth','Standard Studio only','Beeston','',23.10,1,0),
  ('K28','Combi Housing 600 x 720 x 600mm with pan drawer','Standard Studio only','Factory','25 pricelist',89.51,1,0),
  SEC('TALL UNITS'),
  ('K34','ACC Tall Microwave Housing 600 x 2100 x 600mm','DDA rooms only','Factory','25 pricelist',254.50,0,1),
  ('K46','DDA Adjustable Height Mechanism','Rak HA-700-C-BOXY-SINK — floor standing mechanical','Rak Systems','HA-700-C-BOXY-SINK',381.79,0,1),
  SEC('WORKTOPS'),
  ('K40','HPL Worktop 4100 x 650 x 25mm — 0.5 board per kitchen','Egger','Panelco','25 pricelist',73.59,1,1),
  SEC('SPLASHBACKS'),
  ('K52','HPL Splashback — 0.72m2 per standard studio','Egger','Panelco','£65/m2',79.87,1,0),
  ('K52','HPL Splashback — 1.44m2 per ACC studio','Egger','Panelco','£65/m2',159.74,0,1),
  ('K61','Brushed SS Splashback 600 x 600 x 9mm','Studio ×2 | ACC ×1','MPM','25 pricelist',73.73,2,1),
  SEC('INFILLS & END PANELS'),
  ('K64','Wall infill 720 x 70 x 18mm','Studio ×2 | ACC ×1','Factory','25 pricelist',6.41,2,1),
  ('K67','Base infill 720 x 70 x 18mm','Studio ×2 | ACC ×1','Factory','25 pricelist',6.61,2,1),
  ('K70','Base End Panel','ACC only','Factory','25 pricelist',18.42,0,1),
  ('K73','Wall End Panel','ACC only','Factory','25 pricelist',11.60,0,1),
  ('K82','DDA sink modesty panel 1200 x 200 x 18mm','DDA rooms only','Factory','25 pricelist',41.42,0,1),
  SEC('SINKS & TAPS'),
  ('K85','BLANCO Sink and Tap','DINAS 45S MINI + MILA ECO — dry fit only','Blanco','Station Road',175.65,1,1),
  SEC('INSTALLATION (inc. offload, distribution, unpacking, assembly, fixing, waste, snagging)'),
  ('INS','Installation — Studio kitchen','Per studio kitchen','','',427.38,1,0),
  ('INS','Installation — Studio ACC kitchen','Per accessible studio kitchen','','',683.80,0,1),
]

APPLIANCE_ITEMS=[
  SEC('HOBS'),
  ('A10','Hob — 2 ring ceramic cut-off timer with plug','Studios and ACC','Montpellier','MCH29T15',88.85,1,1),
  SEC('EXTRACTORS'),
  ('A25','Extractor — canopy with charcoal filter (CHAR-04)','Studios and ACC','Montpellier','MCA52S',82.11,1,1),
  SEC('MICROWAVES'),
  ('A28','Microwave — integrated combi 45L black','Studios and ACC','Montpellier','MWBIC90044',403.87,1,1),
  SEC('FRIDGE FREEZERS'),
  ('A58','Fridge — undercounter integrated with icebox','Standard Studio','Montpellier','MBUR115E',249.05,1,0),
  ('A61','Fridge — undercounter freestanding with icebox white','Studio ACC','Montpellier','MDAUCIB54W',133.28,0,1),
  SEC('DISTRIBUTION'),
  ('DIS','Distribution of appliances and dry fit','Per appliance — 4 per kitchen','','',12.31,4,4),
]

LOOSE_ITEMS=[
  SEC('MATTRESSES & BEDROOM LOOSE'),
  ('L7','Mattress Double — 1350 x 1900 x 190mm contract','','Stonehouse','Mid matt',88.51,1,1),
  ('L10','Desk Chair — mesh gas lift arms crib 5 black','','TC','Start',73.73,1,1),
  SEC('KITCHEN / COMMUNAL LOOSE'),
  ('L17','Dining Chair — brown or grey','Crib 5, Julian Bowen Brooklyn','Julian Bowen','Brooklyn',80.03,2,2),
  ('L26','Pedestal Table 750 x 750 x 700mm','25mm MFC top, black base','Global','Poland',105.42,1,1),
  SEC('BINS & ACCESSORIES'),
  ('L53','Wastepaper Bin — plastic 12L','','Plastic Box Shop','NWDRB',5.14,1,1),
  ('L59','Integrated Bin — 2 x 10L / 1 x 20L','','Emuca','8199423',49.71,1,1),
  ('L64','Fire blanket 1 x 1m','','Safety Supply Company','SGI-GF-FB-10-10A',5.80,1,1),
  SEC('DISTRIBUTION'),
  ('DIS','Distribution — bins, accessories & brown goods','Per item — approx. 3 per studio','','',3.08,3,3),
  ('DIS','Distribution — assembled goods','Per item — approx. 5 per studio','','',15.39,5,5),
]

ws_bed,bed_tot=build_detail(wb,'BEDROOMS',BEDROOM_ITEMS,STUDIO,ACC_N)
ws_kit,kit_tot=build_detail(wb,'KITCHENS',KITCHEN_ITEMS,STUDIO,ACC_N)
ws_app,app_tot=build_detail(wb,'APPLIANCES',APPLIANCE_ITEMS,STUDIO,ACC_N)
ws_los,los_tot=build_detail(wb,'LOOSE',LOOSE_ITEMS,STUDIO,ACC_N)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET: ROOM SPLITS
# ─────────────────────────────────────────────────────────────────────────────
ws_rs=wb.create_sheet('ROOM SPLITS')
ws_rs.column_dimensions['A'].width=3
for col,w in [('B',22),('C',14),('D',14),('E',14),('F',14),('G',10),('H',14),('I',4)]:
    ws_rs.column_dimensions[col].width=w

ws_rs.merge_cells('B1:H1')
cl=ws_rs['B1']; cl.value='Flagstaffe Room Splits — Heavitree Road Verification'
cl.font=Font(name='Arial',bold=True,size=11,color=WHITE)
cl.fill=fill(NAVY); cl.alignment=al('center','center')
ws_rs.row_dimensions[1].height=20

r=3
hdrs=['Room Type','Total/Room','Bedrooms','Kitchens','Loose','Appliances','Qty','Grand Total']
for i,h in enumerate(hdrs,2):
    cell(ws_rs,r,i,h,NAVY,WHITE,True,'right',sz=9)
    ws_rs.row_dimensions[r].height=15

r+=1
rows_data=[
    ('Standard Studio','£4,290.50','£1,511.07','£1,331.80','£574.52','£873.11',STUDIO,STUDIO*4290.50),
    ('Studio ACC','£5,248.89','£1,734.34','£2,182.68','£574.52','£757.34',ACC_N,ACC_N*5248.89),
]
for rd in rows_data:
    for i,v in enumerate(rd,2):
        bg=LIGHT if i==2 else WHITE
        cell(ws_rs,r,i,v,bg,NAVY if i==2 else '000000',i==2,'right',
             GBP if isinstance(v,float) else (INT if isinstance(v,int) and i==7 else '@'),9)
    ws_rs.row_dimensions[r].height=15; r+=1

# PM Fee row
cell(ws_rs,r,2,'PM Fee (813 units × £68.43)',LIGHT,NAVY,True,'left',sz=9)
ws_rs.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
cell(ws_rs,r,7,TOTAL,LIGHT,NAVY,False,'right',INT,9)
cell(ws_rs,r,8,PM_T,SUBTOT_BG,NAVY,True,'right',GBP,9)
ws_rs.row_dimensions[r].height=15; r+=1

# Sub-total
cell(ws_rs,r,2,'Sub-total',SUBTOT_BG,NAVY,True,'left',sz=10)
ws_rs.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
cell(ws_rs,r,8,SUB_T,SUBTOT_BG,NAVY,True,'right',GBP,10)
ws_rs.row_dimensions[r].height=16; r+=1

# MCD
cell(ws_rs,r,2,'2.5% MCD Discount',LIGHT,RED_F,True,'left',sz=9)
ws_rs.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
mcd_cl=ws_rs.cell(r,8); mcd_cl.value=-MCD_T
mcd_cl.font=Font(name='Arial',bold=True,color=RED_F,size=9)
mcd_cl.fill=fill(LIGHT); mcd_cl.alignment=al('right','center')
mcd_cl.border=bdr(); mcd_cl.number_format=GBP
ws_rs.row_dimensions[r].height=15; r+=1

# Grand total
cell(ws_rs,r,2,'TOTAL (excl. VAT)',NAVY,WHITE,True,'left',sz=12)
ws_rs.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
tc=ws_rs.cell(r,8); tc.value=GRAND
tc.font=Font(name='Arial',bold=True,color=NAVY,size=12)
tc.fill=fill(AMBER); tc.alignment=al('right','center')
tc.border=thick_bdr(); tc.number_format=GBP
ws_rs.row_dimensions[r].height=22; r+=2

# Verify
cell(ws_rs,r,2,'Verification (room totals + PM − MCD should equal TOTAL):',LIGHT,'555555',False,'left',sz=8,italic=True)
r+=1
cell(ws_rs,r,2,'STUDIO × £4,290.50  +  ACC × £5,248.89  +  PM  −  MCD',LIGHT,'555555',False,'left',sz=8)
check=STUDIO*4290.50+ACC_N*5248.89+PM_T-MCD_T
cell(ws_rs,r,8,check,LIGHT,'555555',False,'right',GBP,8)
ws_rs.row_dimensions[r].height=13; r+=1
cell(ws_rs,r,2,'Difference (should show £0.00):',LIGHT,'555555',False,'left',sz=8)
diff=round(GRAND-check,2)
cell(ws_rs,r,8,diff,LIGHT,'2E7D32',True,'right',GBP,8)
ws_rs.row_dimensions[r].height=13

# ─────────────────────────────────────────────────────────────────────────────
# SHEET: T&Cs
# ─────────────────────────────────────────────────────────────────────────────
ws_tc=wb.create_sheet('T&Cs')
ws_tc.column_dimensions['A'].width=3
ws_tc.column_dimensions['B'].width=8
ws_tc.column_dimensions['C'].width=82
ws_tc.column_dimensions['D'].width=4

ws_tc.merge_cells('B1:C1')
cl=ws_tc['B1']; cl.value='TERMS AND CONDITIONS'
cl.font=Font(name='Arial',bold=True,size=12,color=WHITE)
cl.fill=fill(NAVY); cl.alignment=al('center','center')
ws_tc.row_dimensions[1].height=22

tcs=[
  ('Contract',None),
  ('1','Pricing excludes VAT, if applicable.'),
  ('2','The quotation is valid for 60 days for the items listed, quantified and included in the total; anything not mentioned or not included in the total is deemed excluded.'),
  ('3','Models and specification of products quoted may be changed by manufacturers in the course of the project. Flagstaffe Ltd reserves the right to amend the offer accordingly.'),
  ('4','The quotation is based on the DDA requirements as shown on the drawings/specification document. Flagstaffe Ltd takes no responsibility for compliance with DDA regulations.'),
  ('5','The quotation is based on the whole order being placed with Flagstaffe Ltd. Changes to scope may affect pricing.'),
  ('Delivery',None),
  ('1','Skips are excluded (to be provided by Main Contractor); Flagstaffe Ltd will segregate waste as directed by the Main Contractor.'),
  ('2','Horizontal distribution is included; vertical distribution is excluded. Main Contractor to provide goods hoist or forklift with driver.'),
  ('3','Safe access for distribution of materials will be required, with no materials of others impeding access through corridors or rooms.'),
  ('4','Quotation is based on available access for articulated lorries, with off-load within 25 metres of vertical distribution point.'),
  ('5','Minor damage to rooms during distribution and installation, due to the bulky nature of the materials, is to be expected.'),
  ('6','It is recommended that final decorations are carried out after FF&E distribution and installation.'),
  ('7','Builder\'s clean included; sparkle clean excluded.'),
  ('8','All furniture is to be installed in suitable conditions: no standing water, dry walls, temperature 10-30°C, relative humidity 35-65%.'),
  ('9','Unless specifically stated and priced, scribing to uneven floors or walls is not included.'),
  ('Design',None),
  ('1','Flagstaffe Ltd will provide detailed drawings for each room type. Explicit sign-off required before bulk manufacturing.'),
  ('2','Flagstaffe Ltd do not offer wet connections; dry fit included only.'),
  ('3','Sink traps excluded; waste outlet included only.'),
  ('4','Flagstaffe Ltd do not offer electrical connections (other than plugging in, if priced).'),
  ('5','Any commissioning of electrical or wet connections excluded.'),
  ('6','Flagstaffe Ltd will take no responsibility for compliance with Fire Regulations.'),
]

r=3
for num,text in tcs:
    if text is None:
        ws_tc.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
        cell(ws_tc,r,2,num,LIGHT,NAVY,True,'left',sz=10)
        ws_tc.row_dimensions[r].height=16
    else:
        cell(ws_tc,r,2,num,WHITE,NAVY,True,'center',sz=9)
        cell(ws_tc,r,3,text,WHITE,'000000',False,'left',sz=9,wrap=True)
        ws_tc.row_dimensions[r].height=14
    r+=1

r+=1
ws_tc.merge_cells(f'B{r}:C{r}')
cl=ws_tc[f'B{r}']
cl.value='enquiries@flagstaffe.com  |  01260 460 003  |  www.flagstaffe.com  |  CRN: 14154708  |  VAT: 421442927  |  UTR: 2239413036'
cl.font=Font(name='Arial',size=8,color=WHITE)
cl.fill=fill(NAVY); cl.alignment=al('center','center')
ws_tc.row_dimensions[r].height=14

# Reorder sheets
order=['SUMMARY','BEDROOMS','KITCHENS','APPLIANCES','LOOSE','ROOM SPLITS','T&Cs']
for i,name in enumerate(order):
    wb.move_sheet(name,offset=-(wb.index(wb[name])-i))

from openpyxl.worksheet.page import PageMargins
for ws_name in order:
    ws_pg=wb[ws_name]
    ws_pg.page_margins=PageMargins(left=0.5,right=0.5,top=0.75,bottom=0.75)
    ws_pg.print_options.gridLines=False

OUTFILE='/mnt/user-data/outputs/Flag-HVR-001_Heavitree_Full_Bid.xlsx'
wb.save(OUTFILE)
print(f'Saved: {OUTFILE}')
print(f'Sheets: {wb.sheetnames}')
print(f'\nVerification:')
print(f'  Bedrooms:   £{bed_tot:,.2f}  (expected £{BED_T:,.2f})')
print(f'  Kitchens:   £{kit_tot:,.2f}  (expected £{KIT_T:,.2f})')
print(f'  Appliances: £{app_tot:,.2f}  (expected £{APP_T:,.2f})')
print(f'  Loose:      £{los_tot:,.2f}  (expected £{LOOSE_T:,.2f})')
computed=bed_tot+kit_tot+app_tot+los_tot+PM_T-MCD_T
print(f'  Grand total from sheets: £{computed:,.2f}')
print(f'  Expected:                £{GRAND:,.2f}')
print(f'  Match: {abs(computed-GRAND)<1.0}')
